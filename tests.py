"""
This module presents tests for timesheet generation
"""


import openpyxl
from openpyxl.styles import Alignment

import datetime


def write_day(sheet, rst, rstp, colst, colstp, weekday_str):
    """
    Writes days to sheet

    :param sheet: sheet
    :param rst: row start
    :param rstp: row stop
    :param colst: column start
    :param colstp: column stop
    :param weekday_str: string to print
    """
    print('cells for date columns, rows ', colst, colstp, rst, rstp)
    cell = sheet.cell(column=colst, row=rst, value=weekday_str)
    align(cell)

    sheet.merge_cells(start_row=rst, start_column=colst, end_row=rstp, end_column=colstp)


WORKBOOK = openpyxl.Workbook()
WORKSHEET = WORKBOOK.active


def align(cell):
    """
    Alignement
    :param cell: cell
    """
    cell.alignment = Alignment(horizontal='center')


HEADER = 'Timetable'

DATE_START = datetime.datetime.now()

# date_stop

RAWS_HURS = 17
HOURS_MARGIN = 23
GROPUS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
GROPUS_DIVISION = [2, 3, 2, 3]

ROWS_WIDTH = len(GROPUS) * 5

WORKSHEET["A1"] = HEADER
CELL = WORKSHEET["A1"]
align(CELL)

WORKSHEET.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ROWS_WIDTH)

START_DATE = datetime.date(2017, 10, 2)


def write_groups(sheet, row_start, col_start):
    """
    Write grouped cells of students
    :param sheet:
    :param row_start:
    :param col_start:
    """
    print('cells for date columns, rows ', col_start, row_start)
    for ind, group in enumerate(GROPUS_DIVISION):
        print('ind, group', ind, group)
        cell = sheet.cell(column=col_start, row=row_start, value=ind + 1)
        sheet.merge_cells(start_row=row_start,
                          start_column=col_start,
                          end_row=row_start,
                          end_column=col_start + group - 1)

        col_start += group
        align(cell)


def write_groups_alphabet(sheet, row_start, col_start):
    """
    Writes alphabet for gropus
    :param sheet:
    :param row_start:
    :param col_start:
    """
    print('cells for date columns, rows ', col_start, row_start)
    col_start -= 1
    for ind, group in enumerate(GROPUS):
        print('ind, gr', ind, group)

        col_start += 1
        cell = sheet.cell(column=col_start, row=row_start, value=group)
        align(cell)


def write_hours(sheet, row_start, col_start):
    """
    Writing hours
    :param sheet:
    :param row_start:
    :param col_start:
    :return: None
    """
    if col_start > 2:
        print("don't use this not for monday")
        return
    size = 19
    input = '08:00'
    format_string = '%H:%M'

    my_time = datetime.datetime.strptime(input, format_string)

    increment = datetime.timedelta(minutes=60)
    for i in range(size):
        hours = datetime.datetime.strftime(my_time, format_string)
        print(hours)
        my_time += increment
        row_start += 1
        cell = sheet.cell(column=col_start, row=row_start, value=hours)
        align(cell)


def generate_calendar_per_day(work_sheet, day_of_study, start_date):
    """
    Generates calendar per day
    :param work_sheet: Sheet
    :param day_of_std: Day of study f.e. 12.10.2017
    :param start_date: Start date of semester - Must be monday
    """
    print(day_of_study)

    date = day_of_study

    if start_date.weekday() != 0:
        raise Exception("not monday ", start_date.weekday())

    weekday_str = day_of_study.strftime('%A')
    print('current weekday str ', weekday_str)

    date_delta = (day_of_study - start_date).days

    day_of_std = date_delta + 1 - (date_delta // 7 * 2)
    print('day of study', day_of_std)

    row = date_delta // 7 + 2
    print('row', row)

    column = (date_delta % 7) * len(GROPUS) + 2
    print('column', column)

    row_start = (row - 2) * HOURS_MARGIN + 2
    row_stop = (row - 2) * HOURS_MARGIN + 2
    col_start = column
    col_stop = column + len(GROPUS) - 1

    write_day(work_sheet, row_start, row_stop, col_start, col_stop, weekday_str)
    row_start += 1
    row_stop += 1
    write_day(work_sheet, row_start, row_stop, col_start, col_stop, str(date))
    row_start += 1
    write_groups(work_sheet, row_start, col_start)
    row_start += 1
    write_groups_alphabet(work_sheet, row_start, col_start)

    col_start -= 1
    write_hours(work_sheet, row_start, col_start)


for i in range(5):
    day_of_study = datetime.date(2017, 10, 2 + i)
    generate_calendar_per_day(WORKSHEET, day_of_study, start_date=START_DATE)
    day_of_study2 = datetime.date(2017, 10, 9 + i)
    generate_calendar_per_day(WORKSHEET, day_of_study2, start_date=START_DATE)
    day_of_study3 = datetime.date(2017, 10, 16 + i)
    generate_calendar_per_day(WORKSHEET, day_of_study3, start_date=START_DATE)

WORKBOOK.save('test.xlsx')
