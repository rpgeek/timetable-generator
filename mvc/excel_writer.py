"""
This module presents tests for timesheet generation
"""

import datetime
import openpyxl
from models.classtime import Classtime
from mvc.model import Model
from openpyxl.styles import Alignment
from .view import View

class ExcelWriter(View):
    def __init__(self, date_start, day_max, groups, groups_division):
        super().__init__()
        self.__workbook = openpyxl.Workbook()
        self.__worksheet = self.__workbook.active
        self.__header = 'Timetable'

        self.__date_start = date_start
        self.day_end = day_max

        # raw_urs = 17
        self.__hours_margin = 23
        self.__groups = groups
        self.__groups_division = groups_division

        __row_width = len(groups) * 5

        self.__worksheet["A1"] = self.__header
        self.__align(self.__worksheet["A1"])
        self.__worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=__row_width)

        self.__generate_calendar_raw()

    def update(self, file_out):

        self.__save(file_out)

    def __write_day(self, rst, rstp, colst, colstp, weekday_str):
        """
        Writes days to sheet

        :param sheet: sheet
        :param rst: row start
        :param rstp: row stop
        :param colst: column start
        :param colstp: column stop
        :param weekday_str: string to print
        """
        print('cells for date columns, rows ', colst, colstp, rst, rstp)
        cell = self.__worksheet.cell(column=colst, row=rst, value=weekday_str)
        self.__align(cell)
        self.__worksheet.merge_cells(start_row=rst, start_column=colst, end_row=rstp, end_column=colstp)

    def __align(self, cell):
        """
        Alignement
        :param cell: cell
        """
        cell.alignment = Alignment(horizontal='center')

    def __write_groups(self, row_start, col_start):
        """
        Write grouped cells of students
        :param sheet:
        :param row_start:
        :param col_start:
        """
        print('cells for date columns, rows ', col_start, row_start)

        for ind, group in enumerate(self.__groups_division):
            print('ind, group', ind, group)
            cell = self.__worksheet.cell(column=col_start, row=row_start, value=ind + 1)

            self.__worksheet.merge_cells(start_row=row_start,
                                         start_column=col_start,
                                         end_row=row_start,
                                         end_column=col_start + group - 1)

            col_start += group
            self.__align(cell)

    def __write_groups_alphabet(self, row_start, col_start):
        """
        Writes alphabet for gropus
        :param sheet:
        :param row_start:
        :param col_start:
        """
        print('cells for date columns, rows ', col_start, row_start)
        col_start -= 1
        for ind, group in enumerate(self.__groups):
            print('ind, gr', ind, group)

            col_start += 1
            cell = self.__worksheet.cell(column=col_start, row=row_start, value=group)
            self.__align(cell)

    def __write_hours(self, row_start, col_start):
        """
        Writing hours
        :param sheet:
        :param row_start:
        :param col_start:
        :return: None
        """
        if col_start > 2:
            print("don't use this not for monday")
            return
        size = 19
        inpt = '08:00'
        format_string = '%H:%M'

        my_time = datetime.datetime.strptime(inpt, format_string)

        increment = datetime.timedelta(minutes=60)
        for _ in range(size):
            hours = datetime.datetime.strftime(my_time, format_string)
            print(hours)
            my_time += increment
            row_start += 1
            cell = self.__worksheet.cell(column=col_start, row=row_start, value=hours)
            self.__align(cell)

    def __add_cls_for_grp(self, rw_start, cl_start, grps_lst, start_hour, stop_hour, data):
        """
        add classes for group list
        :param sheet: current sheet
        :param rw_start: where to start adding (first row of class)
        :param cl_start: first column
        :param grps_lst: group list - should be 0/1 list of activation
        :param start_hour: start of hour - 1-> 8.00, 2->9.00
        :param stop_hour: as start hour
        :param data: what to put in cell
        """
        col_end = cl_start
        for ind, val in enumerate(grps_lst):

            amount = self.__groups_division[ind]
            print('amount ', amount)
            col_end += amount

            if val:
                self.__worksheet.cell(column=cl_start, row=rw_start, value=data)
                self.__worksheet.merge_cells(start_row=rw_start + start_hour - 1,
                                             start_column=cl_start,
                                             end_row=rw_start + stop_hour - 1,
                                             end_column=col_end - 1)
            else:
                cl_start = col_end

    def add_classtime(self, cls: Classtime, repeated=False):
        self.__add_cls(cls.date, cls.hrs_st, cls.hrs_stp, cls.groups, cls.text, repeated)

    def __add_cls(self, day, start_hour, stop_hour, grp_lst,
                  text, repeated=False, repeat_factor=15):
        """
        Adding classes for day, can be repeatable
        :param day: which date
        :param grp_lst: list of 0/1 grop
        :param start_hour: start hour
        :param stop_hour: stop hor
        :param text: what text to put in cell
        :param repeated: Boolean -> if we want to repead
        :param repeat_factor: how many times repeat
        """
        date_delta = (day - self.__date_start).days
        if repeated:
            for _ in range(repeat_factor):
                day_of_std = date_delta + 1 - (date_delta // 7 * 2)
                print('day of study', day_of_std)

                row = date_delta // 7 + 2
                print('row', row)

                column = (date_delta % 7) * len(self.__groups) + 2
                print('column', column)

                row_start = (row - 2) * self.__hours_margin + 2
                col_start = column
                row_start += 4
                self.__add_cls_for_grp(self.__worksheet, row_start, col_start,
                                       grp_lst, start_hour, stop_hour, text)
                date_delta += 7

        else:
            day_of_std = date_delta + 1 - (date_delta // 7 * 2)
            print('day of study', day_of_std)

            row = date_delta // 7 + 2
            print('row', row)

            column = (date_delta % 7) * len(self.__groups) + 2
            print('column', column)

            row_start = (row - 2) * self.__hours_margin + 2
            col_start = column
            row_start += 4
            self.__add_cls_for_grp(row_start, col_start, grp_lst, start_hour, stop_hour, text)

    def __generate_calendar_per_day(self, day_of_study):
        """
        Generates calendar per day
        :param work_sheet: Sheet
        :param day_of_std: Day of study f.e. 12.10.2017
        :param start_date: Start date of semester - Must be monday
        """
        print(day_of_study)

        date = day_of_study

        if self.__date_start.weekday() != 0:
            raise Exception("not monday ", self.__date_start.weekday())

        weekday_str = day_of_study.strftime('%A')
        print('current weekday str ', weekday_str)

        date_delta = (day_of_study - self.__date_start).days

        day_of_std = date_delta + 1 - (date_delta // 7 * 2)
        print('day of study', day_of_std)

        row = date_delta // 7 + 2
        print('row', row)

        column = (date_delta % 7) * len(self.__groups) + 2
        print('column', column)

        row_start = (row - 2) * self.__hours_margin + 2
        row_stop = (row - 2) * self.__hours_margin + 2
        col_start = column
        col_stop = column + len(self.__groups) - 1

        self.__write_day(row_start, row_stop, col_start, col_stop, weekday_str)
        row_start += 1
        row_stop += 1
        self.__write_day(row_start, row_stop, col_start, col_stop, str(date))
        row_start += 1
        self.__write_groups(row_start, col_start)
        row_start += 1
        self.__write_groups_alphabet(row_start, col_start)

        col_start -= 1
        self.__write_hours(row_start, col_start)

    def __save(self, param):
        self.__workbook.save(param)

    def __generate_calendar_raw(self):

        end = (self.day_end - self.__date_start).days
        print(end)
        j = 0
        while 7 * j < end:
            for i in range(7):
                day_of_study = self.__date_start + datetime.timedelta(days=7 * j + i)
                print('dostudy', day_of_study)
                self.__generate_calendar_per_day(day_of_study)

            j += 1


def test_excel():
    groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    groups_division = [2, 3, 2, 3]
    date_start = datetime.date(2017, 10, 2)

    day_max = datetime.date(2018, 2, 22)

    exwr = ExcelWriter(date_start, day_max, groups, groups_division)
    today = datetime.date.today()

    cls = Classtime('Biologia', today, 1, 4, [1, 0, 0, 1], 'Mr X', '103E')
    exwr.add_classtime(cls)
    exwr.__save("test1.xlsx")
