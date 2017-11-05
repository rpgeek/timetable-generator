"""
This module presents tests for timesheet generation
"""

import datetime
import openpyxl
from openpyxl.styles import Alignment

class ExcelWriter(object):
    def __init__(self):
        self.WORKBOOK = openpyxl.Workbook()
        self.WORKSHEET = self.WORKBOOK.active
        HEADER = 'Timetable'

        DATE_START = datetime.datetime.now()

        # date_stop

        RAWS_HURS = 17
        HOURS_MARGIN = 23
        GROPUS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        GROPUS_DIVISION = [2, 3, 2, 3]

        ROWS_WIDTH = len(GROPUS) * 5

        self.WORKSHEET["A1"] = HEADER
        CELL = self.WORKSHEET["A1"]
        self.align(CELL)

        self.WORKSHEET.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ROWS_WIDTH)

        START_DATE = datetime.date(2017, 10, 2)

    def write_day(self, sheet, rst, rstp, colst, colstp, weekday_str):
        """
        Writes days to sheet

        :param sheet: sheet
        :param rst: row start
        :param rstp: row stop
        :param colst: column start
        :param colstp: column stop
        :param weekday_str: string to print
        """
        print('cells for date columns, rows ', colst, colstp, rst, rstp)
        cell = sheet.cell(column=colst, row=rst, value=weekday_str)
        self.align(cell)

        sheet.merge_cells(start_row=rst, start_column=colst, end_row=rstp, end_column=colstp)

    def align(self, cell):
        """
        Alignement
        :param cell: cell
        """
        cell.alignment = Alignment(horizontal='center')

    def write_groups(self, sheet, row_start, col_start):
        """
        Write grouped cells of students
        :param sheet:
        :param row_start:
        :param col_start:
        """
        print('cells for date columns, rows ', col_start, row_start)
        for ind, group in enumerate(self.GROPUS_DIVISION):
            print('ind, group', ind, group)
            cell = sheet.cell(column=col_start, row=row_start, value=ind + 1)
            sheet.merge_cells(start_row=row_start,
                              start_column=col_start,
                              end_row=row_start,
                              end_column=col_start+group-1)

            col_start += group
            self.align(cell)


    def write_groups_alphabet(self, sheet, row_start, col_start):
        """
        Writes alphabet for gropus
        :param sheet:
        :param row_start:
        :param col_start:
        """
        print('cells for date columns, rows ', col_start, row_start)
        col_start -= 1
        for ind, group in enumerate(self.GROPUS):
            print('ind, gr', ind, group)

            col_start += 1
            cell = sheet.cell(column=col_start, row=row_start, value=group)
            self.align(cell)


    def write_hours(self, sheet, row_start, col_start):
        """
        Writing hours
        :param sheet:
        :param row_start:
        :param col_start:
        :return: None
        """
        if col_start > 2:
            print("don't use this not for monday")
            return
        size = 19
        inpt = '08:00'
        format_string = '%H:%M'

        my_time = datetime.datetime.strptime(inpt, format_string)

        increment = datetime.timedelta(minutes=60)
        for _ in range(size):
            hours = datetime.datetime.strftime(my_time, format_string)
            print(hours)
            my_time += increment
            row_start += 1
            cell = sheet.cell(column=col_start, row=row_start, value=hours)
            self.align(cell)


    def add_cls_for_grp(self, sheet, rw_start, cl_start, grps_lst, start_hour, stop_hour, data):
        """
        add classes for group list
        :param sheet: current sheet
        :param rw_start: where to start adding (first row of class)
        :param cl_start: first column
        :param grps_lst: group list - should be 0/1 list of activation
        :param start_hour: start of hour - 1-> 8.00, 2->9.00
        :param stop_hour: as start hour
        :param data: what to put in cell
        """
        col_end = cl_start
        for ind, val in enumerate(grps_lst):

            amount = self.GROPUS_DIVISION[ind]
            print('amount ', amount)
            col_end += amount

            if val:
                sheet.cell(column=cl_start, row=rw_start, value=data)
                sheet.merge_cells(start_row=rw_start + start_hour - 1,
                                  start_column=cl_start,
                                  end_row=rw_start + stop_hour - 1,
                                  end_column=col_end-1)
            else:
                cl_start = col_end


    def add_classes(self, work_sheet, day, start_date,
                    grp_lst, start_hour, stop_hour,
                    text, repeated=False, repeat_factor=15):

        """
        Adding classes for day, can be repeatable
        :param work_sheet: work sheet
        :param day: which date
        :param start_date: when timetable started
        :param grp_lst: list of 0/1 grop
        :param start_hour: start hour
        :param stop_hour: stop hor
        :param text: what text to put in cell
        :param repeated: Boolean -> if we want to repead
        :param repeat_factor: how many times repeat
        """
        date_delta = (day - start_date).days
        if repeated:
            for _ in range(repeat_factor):
                day_of_std = date_delta + 1 - (date_delta // 7 * 2)
                print('day of study', day_of_std)

                row = date_delta // 7 + 2
                print('row', row)

                column = (date_delta % 7) * len(self.GROPUS) + 2
                print('column', column)

                row_start = (row - 2) * self.HOURS_MARGIN + 2
                col_start = column
                row_start += 4
                self.add_cls_for_grp(work_sheet, row_start, col_start,
                                     grp_lst, start_hour, stop_hour, text)
                date_delta += 7

        else:
            day_of_std = date_delta + 1 - (date_delta // 7 * 2)
            print('day of study', day_of_std)

            row = date_delta // 7 + 2
            print('row', row)

            column = (date_delta % 7) * len(self.GROPUS) + 2
            print('column', column)

            row_start = (row - 2) * self.HOURS_MARGIN + 2
            col_start = column
            row_start += 4
            self.add_cls_for_grp(work_sheet, row_start, col_start, grp_lst, start_hour, stop_hour, text)


    def generate_calendar_per_day(self, work_sheet, day_of_study, start_date):
        """
        Generates calendar per day
        :param work_sheet: Sheet
        :param day_of_std: Day of study f.e. 12.10.2017
        :param start_date: Start date of semester - Must be monday
        """
        print(day_of_study)

        date = day_of_study

        if start_date.weekday() != 0:
            raise Exception("not monday ", start_date.weekday())

        weekday_str = day_of_study.strftime('%A')
        print('current weekday str ', weekday_str)

        date_delta = (day_of_study - start_date).days

        day_of_std = date_delta + 1 - (date_delta // 7 * 2)
        print('day of study', day_of_std)

        row = date_delta // 7 + 2
        print('row', row)

        column = (date_delta % 7) * len(self.GROPUS) + 2
        print('column', column)

        row_start = (row - 2) * self.HOURS_MARGIN + 2
        row_stop = (row - 2) * self.HOURS_MARGIN + 2
        col_start = column
        col_stop = column + len(self.GROPUS) - 1

        self.write_day(work_sheet, row_start, row_stop, col_start, col_stop, weekday_str)
        row_start += 1
        row_stop += 1
        self.write_day(work_sheet, row_start, row_stop, col_start, col_stop, str(date))
        row_start += 1
        self.write_groups(work_sheet, row_start, col_start)
        row_start += 1
        self.write_groups_alphabet(work_sheet, row_start, col_start)

        col_start -= 1
        self.write_hours(work_sheet, row_start, col_start)
#
#
# for i in range(5):
#     day_of_study = datetime.date(2017, 10, 2 + i)
#     generate_calendar_per_day(WORKSHEET, day_of_study, start_date=START_DATE)
#     day_of_study2 = datetime.date(2017, 10, 9 + i)
#     generate_calendar_per_day(WORKSHEET, day_of_study2, start_date=START_DATE)
#     day_of_study3 = datetime.date(2017, 10, 16 + i)
#     generate_calendar_per_day(WORKSHEET, day_of_study3, start_date=START_DATE)
#
#
# add_classes(work_sheet=WORKSHEET, day=datetime.date(2017, 10, 2),
#             start_date=START_DATE, grp_lst=(1, 0, 0, 1),
#             start_hour=1, stop_hour=4,
#             repeated=True, text="Biologia")
#
# WORKBOOK.save('test.xlsx')
